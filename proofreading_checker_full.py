# -*- coding: utf-8 -*-
"""
Full Proofreading Checker for Japanese/English Translation (Bidirectional)
Input: 
 - Bilingual translation file (.txt, tab-separated, 2 columns per line)
 - Glossary file (.xlsx, col A = source language, col B = target language)
Output:
 - Excel file with flagged issues per sentence
   (file name includes detected direction: JP2EN or EN2JP)

Note:
 - Requires the spaCy model 'en_core_web_sm'.
   Install via:  python -m spacy download en_core_web_sm
"""

import re
import spacy
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Alignment
import os
import unicodedata

nlp = spacy.load("en_core_web_sm")


# ========== 1. Language heuristics (similar to ref_sign_checker) ==========

def is_japanese_char(ch: str) -> bool:
    code = ord(ch)
    # Hiragana, Katakana, CJK, full-width forms (rough but sufficient)
    if 0x3040 <= code <= 0x309F:
        return True  # Hiragana
    if 0x30A0 <= code <= 0x30FF:
        return True  # Katakana
    if 0x4E00 <= code <= 0x9FFF:
        return True  # Kanji
    if 0xFF00 <= code <= 0xFFEF:
        return True  # Full-width forms
    return False


def is_english_char(ch: str) -> bool:
    return ("A" <= ch <= "Z") or ("a" <= ch <= "z")


def language_scores_for_text(s: str) -> tuple[int, int]:
    jp = en = 0
    for ch in s:
        if ch.isspace():
            continue
        if is_japanese_char(ch):
            jp += 1
        elif is_english_char(ch):
            en += 1
    return jp, en


def language_scores_for_series(series: pd.Series, max_rows: int = 50) -> tuple[int, int]:
    parts = []
    for val in series.head(max_rows):
        parts.append(str(val))
    text = "".join(parts)
    return language_scores_for_text(text)


def normalize_bilingual_df(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    """
    Given a 2-column DataFrame from a bilingual source, detect which column
    is Japanese and which is English, then return:
      - df with columns ['Japanese', 'English']
      - direction tag 'JP2EN' or 'EN2JP' based on column order
    """
    if df_raw.shape[1] != 2:
        raise ValueError("Expected exactly 2 columns in bilingual data")

    col0, col1 = df_raw.columns
    jp0, en0 = language_scores_for_series(df_raw[col0])
    jp1, en1 = language_scores_for_series(df_raw[col1])

    lang0 = "JP" if jp0 >= en0 else "EN"
    lang1 = "JP" if jp1 >= en1 else "EN"

    if lang0 == "JP" and lang1 == "EN":
        jp_col, en_col = col0, col1
    elif lang0 == "EN" and lang1 == "JP":
        jp_col, en_col = col1, col0
    else:
        # Fallback: assume original order [JP, EN]
        jp_col, en_col = col0, col1

    # Match ref_sign_checker style: JP2EN if JP column index < EN column index
    jp_index = 0 if jp_col == col0 else 1
    en_index = 1 - jp_index
    direction = "JP2EN" if jp_index < en_index else "EN2JP"

    df = pd.DataFrame(
        {
            "Japanese": df_raw[jp_col],
            "English": df_raw[en_col],
        }
    )
    return df, direction


# ========== 2. Loaders & normalizers ==========

def load_bilingual_txt(path: str) -> pd.DataFrame:
    """
    Load a tab-separated bilingual TXT and normalize to ['Japanese', 'English'].
    Direction detection is done but ignored here (mainly for CLI use).
    """
    lines = Path(path).read_text(encoding="utf-8").splitlines()
    pairs = [tuple(line.split("\t")) for line in lines if "\t" in line]
    df_raw = pd.DataFrame(pairs, columns=["col0", "col1"])
    df, _ = normalize_bilingual_df(df_raw)
    return df


def normalize_en(s: str) -> str:
    return str(s).lower()


def normalize_ja(s: str) -> str:
    # NFKC to unify full-/half-width, but DO NOT remove spaces
    s = unicodedata.normalize("NFKC", str(s))
    # Convert full-width space to normal space (keep spaces)
    s = s.replace("\u3000", " ")
    return s


def split_variants(s: str) -> list[str]:
    # Split on half- or full-width semicolon
    return [v.strip() for v in re.split(r"[;；]", str(s)) if v.strip()]


def load_glossary_xlsx(path: str) -> tuple[list[tuple[str, list[str]]], str]:
    """
    Load a glossary Excel file and detect its direction:

      - Column A: source language terms (JP or EN)
      - Column B: target language terms (variants separated by ';' or '；')

    Direction detection:
      - If col A looks JP and col B looks EN → 'JP2EN'
      - If col A looks EN and col B looks JP → 'EN2JP'
      - Otherwise default to 'JP2EN'

    Returns:
      glossary: list of (src_term, [tgt_variants])
      direction: 'JP2EN' or 'EN2JP'
    """
    df = pd.read_excel(path)
    if df.shape[1] < 2:
        raise ValueError("Glossary must have at least two columns (source, target)")

    col0 = df.iloc[:, 0]
    col1 = df.iloc[:, 1]

    jp0, en0 = language_scores_for_series(col0)
    jp1, en1 = language_scores_for_series(col1)

    lang0 = "JP" if jp0 >= en0 else "EN"
    lang1 = "JP" if jp1 >= en1 else "EN"

    if lang0 == "JP" and lang1 == "EN":
        direction = "JP2EN"
        src_series = col0
        tgt_series = col1
    elif lang0 == "EN" and lang1 == "JP":
        direction = "EN2JP"
        src_series = col0
        tgt_series = col1
    else:
        # Fallback: assume JP→EN
        direction = "JP2EN"
        src_series = col0
        tgt_series = col1

    glossary: list[tuple[str, list[str]]] = []
    for src, tgt in zip(src_series, tgt_series):
        if pd.isna(src) or pd.isna(tgt):
            continue
        src_str = str(src)
        tgt_variants = split_variants(tgt)
        if src_str and tgt_variants:
            glossary.append((src_str, tgt_variants))

    return glossary, direction


# ========== 3. Core English-side checks ==========

def check_word_repetition(sentence: str) -> list[str]:
    """
    Detect only *strict* consecutive repetition in English:
      e.g., 'policy policy', 'the   the'
    Ignore cases like 'policy (policy', 'policy, policy', etc.
    """
    if not isinstance(sentence, str):
        return []

    doc = nlp(sentence)
    issues: list[str] = []

    for i in range(len(doc) - 1):
        t1 = doc[i]
        t2 = doc[i + 1]

        if not (t1.is_alpha and t2.is_alpha):
            continue

        if t1.text.lower() != t2.text.lower():
            continue

        between = sentence[t1.idx + len(t1.text) : t2.idx]
        if between.strip() != "":
            continue

        issues.append(f"Consecutive repetition: '{t1.text} {t2.text}'")

    return issues


def check_double_space(sentence: str) -> list[str]:
    """
    Detect runs of two or more consecutive half-width spaces in English.
    - Flags "  ", "   ", "    ", etc.
    - Each run is reported once.
    """
    if not isinstance(sentence, str):
        return []

    issues: list[str] = []
    pattern = re.compile(r" {2,}")

    for m in pattern.finditer(sentence):
        run_len = len(m.group())
        idx = m.start()
        context = sentence[max(0, idx - 10) : idx + run_len + 10].replace("\n", "\\n")
        shown = " " * min(run_len, 4)
        issues.append(
            f"Consecutive spaces ({run_len}): '{shown}' (context: …{context}…)"
        )

    return issues


# ========== 4. Bidirectional glossary checker ==========

def check_glossary_terms(
    src_text: str,
    tgt_text: str,
    glossary: list[tuple[str, list[str]]],
    src_lang: str,
) -> list[str]:
    """
    Generic glossary checker.

    glossary: list of (src_term, [tgt_variant1, ...])
    src_lang: 'JP' or 'EN' (language of src_term and src_text)

    Behavior:
      - If normalized src_term appears in normalized src_text,
        require at least one normalized tgt_variant in normalized tgt_text.
    """
    flags: list[str] = []

    if not isinstance(src_text, str) or not isinstance(tgt_text, str):
        return flags

    if src_lang == "JP":
        src_text_norm = normalize_ja(src_text)
        tgt_text_norm = normalize_en(tgt_text)
        for src_term, tgt_variants in glossary:
            src_term_norm = normalize_ja(src_term)
            if src_term_norm and src_term_norm in src_text_norm:
                if not any(
                    normalize_en(v) in tgt_text_norm for v in tgt_variants
                ):
                    flags.append(
                        f"Glossary missing: '{src_term}' → {', '.join(tgt_variants)}"
                    )
    else:  # src_lang == "EN"
        src_text_norm = normalize_en(src_text)
        tgt_text_norm = normalize_ja(tgt_text)
        for src_term, tgt_variants in glossary:
            src_term_norm = normalize_en(src_term)
            if src_term_norm and src_term_norm in src_text_norm:
                if not any(
                    normalize_ja(v) in tgt_text_norm for v in tgt_variants
                ):
                    flags.append(
                        f"Glossary missing: '{src_term}' → {', '.join(tgt_variants)}"
                    )

    return flags


# ========== 5. Master runner ==========

def run_all_checks(
    df: pd.DataFrame,
    glossary: list[tuple[str, list[str]]],
    glossary_direction: str,
) -> pd.DataFrame:
    """
    df: DataFrame with columns ['Japanese', 'English']
    glossary: list of (src_term, [tgt_variants])
    glossary_direction: 'JP2EN' or 'EN2JP'
    """
    results = []
    src_lang = "JP" if glossary_direction == "JP2EN" else "EN"

    for _, row in df.iterrows():
        jp = row["Japanese"]
        en = row["English"]

        flags: list[str] = []

        # English-side form checks always on English column
        flags += check_word_repetition(en)
        flags += check_double_space(en)

        # Glossary checks (directional)
        if src_lang == "JP":
            flags += check_glossary_terms(jp, en, glossary, src_lang="JP")
        else:
            flags += check_glossary_terms(en, jp, glossary, src_lang="EN")

        results.append(
            {"Japanese": jp, "English": en, "Issues": "; ".join(flags)}
        )

    return pd.DataFrame(results)


# ========== 6. Excel export (direction in file name) ==========

def export_with_format(
    out_df: pd.DataFrame,
    src_path: str,
    direction_tag: str
) -> str:
    """
    Export results to a single Excel file with two sheets:
      - 'All'
      - 'IssuesOnly'

    Column order:
      - JP2EN: A=Japanese, B=English
      - EN2JP: A=English,  B=Japanese
    """

    # Decide column order based on direction
    if direction_tag == "EN2JP":
        cols = ["English", "Japanese", "Issues"]
    else:  # JP2EN
        cols = ["Japanese", "English", "Issues"]

    all_df = out_df[cols]

    issue_mask = all_df["Issues"].astype(str).str.strip() != ""
    issues_df = all_df[issue_mask]

    out_dir = os.path.dirname(src_path)
    stamp = datetime.now().strftime("%Y%m%d")
    out_file = os.path.join(
        out_dir,
        f"proofreading_result_{direction_tag}_{stamp}.xlsx"
    )

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        all_df.to_excel(writer, index=False, sheet_name="All")
        issues_df.to_excel(writer, index=False, sheet_name="IssuesOnly")

        book = writer.book
        for sheet in book.worksheets:
            sheet.column_dimensions["A"].width = 60
            sheet.column_dimensions["B"].width = 60
            sheet.column_dimensions["C"].width = 60

            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=1,
                max_col=sheet.max_column,
            ):
                for cell in row:
                    if cell.row == 1 and cell.column in (1, 2, 3):
                        cell.alignment = Alignment(
                            wrapText=True,
                            horizontal="center"
                        )
                    else:
                        cell.alignment = Alignment(wrapText=True)

    return out_file


# ========== 7. CLI entry (for TXT input) ==========

def main():
    txt = input("Enter path of bilingual text: ").strip('"')
    df = load_bilingual_txt(txt)

    term_list_file = input("Enter path of term list: ").strip('"')
    glossary, glossary_direction = load_glossary_xlsx(term_list_file)
    print(f"Detected glossary direction: {glossary_direction}")

    out_df = run_all_checks(df, glossary, glossary_direction)
    out_file = export_with_format(out_df, txt, glossary_direction)
    print(f"Saved: {out_file}")


if __name__ == "__main__":
    main()
